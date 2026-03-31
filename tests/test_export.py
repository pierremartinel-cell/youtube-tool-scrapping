import io

import openpyxl
import pytest

from youtube_scraper import COLUMNS, calculate_tier, compute_scores, export_excel


@pytest.fixture
def mock_profiles():
    profiles = []
    for i in range(10):
        followers = (i + 1) * 15_000
        eng = 0.01 + i * 0.008
        mentions = i % 5
        ppw = 0.5 + i * 0.3
        growth = i * 2.5
        se, sc, sp, sr, sg = compute_scores(eng, mentions, ppw, growth, has_video_stats=True)
        profiles.append(
            {
                "platform": "YouTube",
                "username": f"creator_{i}",
                "display_name": f"Test Creator {i}",
                "profile_url": f"https://www.youtube.com/@creator_{i}",
                "email": f"creator{i}@example.com",
                "bio_snippet": f"Bio of creator {i}.",
                "followers": followers,
                "tier": calculate_tier(followers),
                "engagement_rate": round(eng, 6),
                "engagement_rate_pct": round(eng * 100, 3),
                "croissance_hebdo": round(followers / 52, 1),
                "growth_rate_pct": round(growth, 2),
                "posts_per_week": round(ppw, 2),
                "sorare_mentions": mentions,
                "is_emerging": growth > 5 and followers < 50_000,
                "score_global": sg,
                "score_engagement": se,
                "score_croissance": sc,
                "score_pertinence": sp,
                "score_regularite": sr,
                "total_recent_views": (i + 1) * 1_000,
                "total_recent_likes": (i + 1) * 50,
                "total_recent_comments": (i + 1) * 5,
                "recent_video_count": (i + 1) * 3,
                "status": "active" if ppw >= 0.5 else "inactive",
                "collected_at": "2025-01-01 12:00:00",
            }
        )
    return profiles


class TestExportExcel:
    def test_export_to_buffer(self, mock_profiles):
        buf = io.BytesIO()
        export_excel(mock_profiles, buf, ["test"])
        buf.seek(0)
        assert len(buf.read()) > 1000

    def test_export_to_file(self, tmp_path, mock_profiles):
        filepath = tmp_path / "test_output.xlsx"
        export_excel(mock_profiles, str(filepath), ["test"])
        assert filepath.exists()
        assert filepath.stat().st_size > 1000

    def test_sheet_names(self, tmp_path, mock_profiles):
        filepath = tmp_path / "test_output.xlsx"
        export_excel(mock_profiles, str(filepath), ["test"])
        wb = openpyxl.load_workbook(filepath)
        assert "Profiles" in wb.sheetnames
        assert "Summary" in wb.sheetnames

    def test_column_headers(self, tmp_path, mock_profiles):
        filepath = tmp_path / "test_output.xlsx"
        export_excel(mock_profiles, str(filepath), ["test"])
        wb = openpyxl.load_workbook(filepath)
        ws = wb["Profiles"]
        headers = [ws.cell(1, c).value for c in range(1, len(COLUMNS) + 1)]
        assert headers == COLUMNS

    def test_row_count(self, tmp_path, mock_profiles):
        filepath = tmp_path / "test_output.xlsx"
        export_excel(mock_profiles, str(filepath), ["test"])
        wb = openpyxl.load_workbook(filepath)
        ws = wb["Profiles"]
        data_rows = ws.max_row - 1  # minus header
        assert data_rows == len(mock_profiles)
