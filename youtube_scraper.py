#!/usr/bin/env python3
"""
YouTube Profile Scraper
========================
Fetches YouTube channel profiles based on keywords, country, and time period.
Outputs an Excel file with scoring indices.

Usage:
    python youtube_scraper.py --keywords "Sorare" --region FR --days 90
    python youtube_scraper.py --keywords "Sorare" "NFT" --region FR --days 90 --output results.xlsx
    python youtube_scraper.py --keywords "Sorare" --region FR --days 90 --max-channels 200
"""

import argparse
import logging
import os
import re
import sys
import time
from datetime import UTC, datetime, timedelta

import pandas as pd
from dotenv import load_dotenv
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tqdm import tqdm

load_dotenv()

logger = logging.getLogger(__name__)

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

TIER_BOUNDARIES = [(1_000_000, "mega"), (100_000, "macro"), (10_000, "mid"), (1_000, "micro"), (0, "nano")]
SCORE_WEIGHTS = {"pertinence": 0.37, "engagement": 0.28, "croissance": 0.20, "regularite": 0.15}
ENGAGEMENT_THRESHOLDS = [(0.10, 100), (0.07, 85), (0.05, 70), (0.03, 55), (0.01, 35), (0, 15)]
PERTINENCE_THRESHOLDS = [(10, 100), (5, 85), (3, 65), (2, 45), (1, 25), (0, 0)]
REGULARITE_THRESHOLDS = [(4, 100), (3, 85), (2, 70), (1, 50), (0.5, 30), (0, 10)]
CROISSANCE_THRESHOLDS = [(30, 100), (20, 85), (10, 65), (5, 45), (1, 25), (0, 10)]
EMERGING_GROWTH_MIN_PCT = 5
EMERGING_FOLLOWERS_MAX = 50_000
RATE_LIMIT_SEARCH = 0.2
RATE_LIMIT_CHANNEL_DETAILS = 0.1
RATE_LIMIT_VIDEO_STATS = 0.15
ACTIVE_THRESHOLD_PPW = 0.5
ZERO_VIDEO_STATS = {"views": 0, "likes": 0, "comments": 0, "video_count": 0}
FAST_MODE_BATCH_SIZE = 50


# ---------------------------------------------------------------------------
# YouTube API client
# ---------------------------------------------------------------------------


def get_youtube_client(api_key: str):
    return build("youtube", "v3", developerKey=api_key, cache_discovery=False)


# ---------------------------------------------------------------------------
# Data collection
# ---------------------------------------------------------------------------


def search_videos_by_keyword(
    youtube,
    keyword: str,
    region_code: str | None,
    days: int,
    language: str | None = None,
    max_channels: int = 150,
) -> dict[str, dict]:
    """
    Search YouTube videos by keyword + region + recency.
    region_code=None means worldwide (no region filter).
    Returns a dict keyed by channel_id with mention counts and video ids.
    """
    published_after = (datetime.now(UTC) - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")

    channels: dict[str, dict] = {}
    next_page_token = None

    while True:
        try:
            params = dict(
                q=keyword,
                type="video",
                publishedAfter=published_after,
                part="snippet",
                maxResults=50,
                order="relevance",
            )
            if region_code:
                params["regionCode"] = region_code
            if language:
                params["relevanceLanguage"] = language
            if next_page_token:
                params["pageToken"] = next_page_token

            response = youtube.search().list(**params).execute()

        except HttpError:
            raise  # propagate to caller for proper UI error handling

        for item in response.get("items", []):
            channel_id = item["snippet"]["channelId"]
            title = item["snippet"].get("title", "")
            description = item["snippet"].get("description", "")
            video_id = item["id"].get("videoId", "")

            if channel_id not in channels:
                channels[channel_id] = {
                    "channel_id": channel_id,
                    "display_name": item["snippet"]["channelTitle"],
                    "video_ids": [],
                    "mentions_count": 0,
                }

            # Count keyword mentions in title or description
            kw_lower = keyword.lower()
            if kw_lower in title.lower() or kw_lower in description.lower():
                channels[channel_id]["mentions_count"] += 1

            if video_id:
                channels[channel_id]["video_ids"].append(video_id)

        next_page_token = response.get("nextPageToken")
        if not next_page_token or len(channels) >= max_channels:
            break

        time.sleep(RATE_LIMIT_SEARCH)

    return channels


def get_channel_details(youtube, channel_ids: list[str]) -> dict[str, dict]:
    """Fetch channel metadata and statistics in batches of 50."""
    result = {}

    for i in range(0, len(channel_ids), 50):
        batch = channel_ids[i : i + 50]
        try:
            response = youtube.channels().list(id=",".join(batch), part="snippet,statistics").execute()
        except HttpError:
            raise  # propagate to caller

        for item in response.get("items", []):
            cid = item["id"]
            stats = item.get("statistics", {})
            snippet = item.get("snippet", {})

            subscribers = stats.get("subscriberCount")
            full_description = snippet.get("description", "")
            email_match = EMAIL_RE.search(full_description)
            result[cid] = {
                "username": snippet.get("customUrl", "").lstrip("@"),
                "display_name": snippet.get("title", ""),
                "bio_snippet": full_description[:250].replace("\n", " "),
                "email": email_match.group(0) if email_match else "",
                "country": snippet.get("country", ""),
                "published_at": snippet.get("publishedAt", ""),
                "followers": int(subscribers) if subscribers else 0,
                "total_views": int(stats.get("viewCount", 0) or 0),
                "total_video_count": int(stats.get("videoCount", 0) or 0),
                "hidden_subscribers": stats.get("hiddenSubscriberCount", False),
            }

        time.sleep(RATE_LIMIT_CHANNEL_DETAILS)

    return result


def get_recent_video_stats(youtube, channel_id: str, days: int) -> dict:
    """Fetch aggregate stats (views, likes, comments) for recent videos."""
    published_after = (datetime.now(UTC) - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")

    try:
        search_resp = (
            youtube.search()
            .list(
                channelId=channel_id,
                type="video",
                publishedAfter=published_after,
                part="id",
                maxResults=50,
                order="date",
            )
            .execute()
        )
    except HttpError:
        logger.warning("Failed to search videos for channel %s", channel_id)
        return dict(ZERO_VIDEO_STATS)

    video_ids = [item["id"]["videoId"] for item in search_resp.get("items", []) if item["id"].get("videoId")]

    if not video_ids:
        return dict(ZERO_VIDEO_STATS)

    try:
        stats_resp = youtube.videos().list(id=",".join(video_ids), part="statistics").execute()
    except HttpError:
        logger.warning("Failed to fetch video stats for channel %s", channel_id)
        return {"views": 0, "likes": 0, "comments": 0, "video_count": len(video_ids)}

    total_views = total_likes = total_comments = 0
    for item in stats_resp.get("items", []):
        s = item.get("statistics", {})
        total_views += int(s.get("viewCount", 0) or 0)
        total_likes += int(s.get("likeCount", 0) or 0)
        total_comments += int(s.get("commentCount", 0) or 0)

    return {
        "views": total_views,
        "likes": total_likes,
        "comments": total_comments,
        "video_count": len(video_ids),
    }


def get_video_stats_batch(youtube, video_ids: list[str]) -> dict:
    """Fetch aggregate stats for a list of video IDs (from keyword search).

    Uses videos().list() directly — no per-channel search needed.
    Batches in groups of FAST_MODE_BATCH_SIZE (50). Costs 1 unit per batch
    instead of 100 units per channel search.
    """
    if not video_ids:
        return dict(ZERO_VIDEO_STATS)

    total_views = total_likes = total_comments = 0
    fetched_count = 0

    for i in range(0, len(video_ids), FAST_MODE_BATCH_SIZE):
        batch = video_ids[i : i + FAST_MODE_BATCH_SIZE]
        try:
            resp = youtube.videos().list(id=",".join(batch), part="statistics").execute()
        except HttpError:
            logger.warning("Failed to fetch video stats batch (offset %d)", i)
            continue

        for item in resp.get("items", []):
            s = item.get("statistics", {})
            total_views += int(s.get("viewCount", 0) or 0)
            total_likes += int(s.get("likeCount", 0) or 0)
            total_comments += int(s.get("commentCount", 0) or 0)
            fetched_count += 1

        time.sleep(RATE_LIMIT_VIDEO_STATS)

    return {
        "views": total_views,
        "likes": total_likes,
        "comments": total_comments,
        "video_count": fetched_count,
    }


# ---------------------------------------------------------------------------
# Scoring helpers
# ---------------------------------------------------------------------------


def calculate_tier(followers: int) -> str:
    for boundary, tier in TIER_BOUNDARIES:
        if followers >= boundary:
            return tier
    return "nano"


def _score_from_thresholds(value: float, thresholds: list[tuple[float, float]], default: float = 0) -> float:
    for min_value, score in thresholds:
        if value >= min_value:
            return score
    return default


def score_engagement(rate: float) -> float:
    """Rate is a ratio (e.g. 0.05 = 5%)."""
    return _score_from_thresholds(rate, ENGAGEMENT_THRESHOLDS, default=15)


def score_pertinence(mentions: int) -> float:
    return _score_from_thresholds(mentions, PERTINENCE_THRESHOLDS, default=0)


def score_regularite(posts_per_week: float) -> float:
    return _score_from_thresholds(posts_per_week, REGULARITE_THRESHOLDS, default=10)


def score_croissance(growth_rate_pct: float) -> float:
    return _score_from_thresholds(growth_rate_pct, CROISSANCE_THRESHOLDS, default=10)


def compute_scores(engagement_rate, mentions, posts_per_week, growth_rate_pct, has_video_stats=False):
    w = SCORE_WEIGHTS
    sp = score_pertinence(mentions)
    sr = score_regularite(posts_per_week)
    sc = score_croissance(growth_rate_pct)

    if has_video_stats:
        se = score_engagement(engagement_rate)
        sg = round(se * w["engagement"] + sp * w["pertinence"] + sr * w["regularite"] + sc * w["croissance"], 1)
    else:
        se = 0
        non_engagement = w["pertinence"] + w["croissance"] + w["regularite"]
        sg = round(
            sp * (w["pertinence"] / non_engagement)
            + sc * (w["croissance"] / non_engagement)
            + sr * (w["regularite"] / non_engagement),
            1,
        )

    return round(se, 1), round(sc, 1), round(sp, 1), round(sr, 1), sg


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

COLUMNS = [
    "platform",
    "username",
    "display_name",
    "profile_url",
    "email",
    "bio_snippet",
    "followers",
    "tier",
    "engagement_rate",
    "engagement_rate_pct",
    "croissance_hebdo",
    "growth_rate_pct",
    "posts_per_week",
    "sorare_mentions",
    "is_emerging",
    "score_global",
    "score_engagement",
    "score_croissance",
    "score_pertinence",
    "score_regularite",
    "total_recent_views",
    "total_recent_likes",
    "total_recent_comments",
    "recent_video_count",
    "status",
    "collected_at",
]

SCORE_COLS = ["score_global", "score_engagement", "score_croissance", "score_pertinence", "score_regularite"]

TIER_COLORS = {
    "mega": "7B2FBE",
    "macro": "3B82F6",
    "mid": "10B981",
    "micro": "F59E0B",
    "nano": "6B7280",
}

HEADER_BG = "1E3A5F"
ALT_ROW_BG = "EEF4FB"


def export_excel(profiles: list[dict], output_file, keywords: list[str]):
    df = pd.DataFrame(profiles, columns=COLUMNS)
    df.sort_values("score_global", ascending=False, inplace=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Profiles")

        wb = writer.book
        ws = writer.sheets["Profiles"]

        # ---- Header styling ----
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_fill = PatternFill("solid", fgColor=HEADER_BG)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.row_dimensions[1].height = 32

        # ---- Column widths ----
        col_widths = {
            "platform": 12,
            "username": 22,
            "display_name": 28,
            "profile_url": 40,
            "email": 30,
            "bio_snippet": 45,
            "followers": 14,
            "tier": 10,
            "engagement_rate": 18,
            "engagement_rate_pct": 20,
            "croissance_hebdo": 18,
            "growth_rate_pct": 18,
            "posts_per_week": 16,
            "sorare_mentions": 18,
            "is_emerging": 14,
            "score_global": 14,
            "score_engagement": 18,
            "score_croissance": 18,
            "score_pertinence": 18,
            "score_regularite": 18,
            "total_recent_views": 18,
            "total_recent_likes": 18,
            "total_recent_comments": 20,
            "recent_video_count": 18,
            "status": 12,
            "collected_at": 20,
        }
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 16)

        # ---- Row styling ----
        col_map = {name: idx + 1 for idx, name in enumerate(COLUMNS)}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            # Alternating row background
            bg = ALT_ROW_BG if row_idx % 2 == 0 else "FFFFFF"
            alt_fill = PatternFill("solid", fgColor=bg)

            for cell in row:
                cell.fill = alt_fill
                cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Tier badge color
            tier_cell = ws.cell(row=row_idx, column=col_map["tier"])
            tier_val = tier_cell.value or ""
            color = TIER_COLORS.get(tier_val, "6B7280")
            tier_cell.font = Font(color=color, bold=True)

            # Hyperlink on display_name (nom cliquable → chaîne YouTube)
            name_cell = ws.cell(row=row_idx, column=col_map["display_name"])
            url_cell = ws.cell(row=row_idx, column=col_map["profile_url"])
            if url_cell.value:
                name_cell.hyperlink = url_cell.value
                name_cell.font = Font(color="1155CC", underline="single", bold=True)

            # Hyperlink on profile_url également
            if url_cell.value:
                url_cell.hyperlink = url_cell.value
                url_cell.font = Font(color="1155CC", underline="single")

            # is_emerging boolean → yes/no
            em_cell = ws.cell(row=row_idx, column=col_map["is_emerging"])
            em_cell.value = "Yes" if em_cell.value else "No"
            if em_cell.value == "Yes":
                em_cell.font = Font(color="10B981", bold=True)

        # ---- Color scale on score columns ----
        for score_col in SCORE_COLS:
            col_letter = get_column_letter(col_map[score_col])
            cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
            rule = ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="FF4444",
                mid_type="num",
                mid_value=50,
                mid_color="FFAA00",
                end_type="num",
                end_value=100,
                end_color="00CC44",
            )
            ws.conditional_formatting.add(cell_range, rule)

        # ---- Summary sheet ----
        ws_sum = wb.create_sheet("Summary")
        ws_sum["A1"] = "YouTube Scraper — Summary"
        ws_sum["A1"].font = Font(bold=True, size=14, color=HEADER_BG)

        summary_data = [
            ("Generated at", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Keywords", ", ".join(keywords)),
            ("Total profiles", len(profiles)),
            ("", ""),
            ("Tier breakdown", ""),
        ]
        tier_counts = df["tier"].value_counts().to_dict()
        for tier in ["mega", "macro", "mid", "micro", "nano"]:
            summary_data.append((f"  {tier}", tier_counts.get(tier, 0)))

        summary_data += [
            ("", ""),
            ("Avg score_global", round(df["score_global"].mean(), 1)),
            ("Avg engagement_rate_pct", round(df["engagement_rate_pct"].mean(), 2)),
            ("Channels with mentions > 0", int((df["sorare_mentions"] > 0).sum())),
            ("Emerging channels", int((df["is_emerging"] == "Yes").sum())),
        ]

        for r, (label, value) in enumerate(summary_data, start=3):
            ws_sum.cell(row=r, column=1, value=label).font = Font(bold=bool(label and not label.startswith(" ")))
            ws_sum.cell(row=r, column=2, value=value)

        ws_sum.column_dimensions["A"].width = 30
        ws_sum.column_dimensions["B"].width = 25

    logger.info("Saved → %s", output_file)


# ---------------------------------------------------------------------------
# Shared logic (used by both scrape() and app.py)
# ---------------------------------------------------------------------------


def merge_keyword_results(all_channels: dict[str, dict], new_channels: dict[str, dict]) -> None:
    """Merge new keyword search results into all_channels, deduplicating by channel id and video id."""
    for cid, data in new_channels.items():
        if cid not in all_channels:
            all_channels[cid] = data
        else:
            all_channels[cid]["mentions_count"] += data["mentions_count"]
            existing_ids = set(all_channels[cid]["video_ids"])
            for vid in data["video_ids"]:
                if vid not in existing_ids:
                    all_channels[cid]["video_ids"].append(vid)
                    existing_ids.add(vid)


def compute_channel_metrics(details: dict, vstats: dict, search_data: dict, days: int) -> dict:
    """Compute engagement rate, posts/week, growth, emerging flag from raw data."""
    followers = details.get("followers", 0)
    total_views = vstats["views"]
    total_likes = vstats["likes"]
    total_comments = vstats["comments"]
    video_count = vstats["video_count"]

    engagement_rate = (total_likes + total_comments) / total_views if total_views > 0 else 0.0
    weeks = days / 7
    posts_per_week = round(video_count / weeks, 2) if weeks > 0 else 0

    published_at_str = details.get("published_at", "")
    if published_at_str:
        try:
            created = datetime.fromisoformat(published_at_str.replace("Z", "+00:00"))
            age_weeks = max((datetime.now(UTC) - created).days / 7, 1)
            croissance_hebdo = round(followers / age_weeks, 1)
            growth_rate_pct = round((croissance_hebdo / max(followers, 1)) * 100, 2)
        except ValueError:
            croissance_hebdo = 0.0
            growth_rate_pct = 0.0
    else:
        croissance_hebdo = 0.0
        growth_rate_pct = 0.0

    mentions = search_data.get("mentions_count", 0)
    is_emerging = growth_rate_pct > EMERGING_GROWTH_MIN_PCT and followers < EMERGING_FOLLOWERS_MAX

    return {
        "engagement_rate": engagement_rate,
        "posts_per_week": posts_per_week,
        "croissance_hebdo": croissance_hebdo,
        "growth_rate_pct": growth_rate_pct,
        "mentions": mentions,
        "is_emerging": is_emerging,
        "followers": followers,
        "total_recent_views": total_views,
        "total_recent_likes": total_likes,
        "total_recent_comments": total_comments,
        "recent_video_count": video_count,
    }


def build_channel_profile(
    cid: str,
    details: dict,
    search_data: dict,
    metrics: dict,
    has_video_stats: bool,
    collected_at: str,
) -> dict:
    """Build a full profile dict with correct scoring, status, and email field."""
    m = metrics
    se, sc, sp, sr, sg = compute_scores(
        m["engagement_rate"],
        m["mentions"],
        m["posts_per_week"],
        m["growth_rate_pct"],
        has_video_stats=has_video_stats,
    )

    username = details.get("username") or cid
    display_name = details.get("display_name") or search_data.get("display_name", "")

    if details.get("username"):
        profile_url = f"https://www.youtube.com/@{details['username'].lstrip('@')}"
    else:
        profile_url = f"https://www.youtube.com/channel/{cid}"

    # Status: when video stats are not fetched, we can't know posting frequency — default to "active"
    if has_video_stats:
        status = "active" if m["posts_per_week"] >= ACTIVE_THRESHOLD_PPW else "inactive"
    else:
        status = "active"

    return {
        "platform": "YouTube",
        "username": username,
        "display_name": display_name,
        "profile_url": profile_url,
        "email": details.get("email", ""),
        "bio_snippet": details.get("bio_snippet", ""),
        "followers": m["followers"],
        "tier": calculate_tier(m["followers"]),
        "engagement_rate": round(m["engagement_rate"], 6),
        "engagement_rate_pct": round(m["engagement_rate"] * 100, 3),
        "croissance_hebdo": m["croissance_hebdo"],
        "growth_rate_pct": m["growth_rate_pct"],
        "posts_per_week": m["posts_per_week"],
        "sorare_mentions": m["mentions"],
        "is_emerging": m["is_emerging"],
        "score_global": sg,
        "score_engagement": se,
        "score_croissance": sc,
        "score_pertinence": sp,
        "score_regularite": sr,
        "total_recent_views": m["total_recent_views"],
        "total_recent_likes": m["total_recent_likes"],
        "total_recent_comments": m["total_recent_comments"],
        "recent_video_count": m["recent_video_count"],
        "status": status,
        "collected_at": collected_at,
    }


# ---------------------------------------------------------------------------
# Main orchestration
# ---------------------------------------------------------------------------


def scrape(
    keywords: list[str],
    region_code: str = "FR",
    days: int = 90,
    language: str | None = None,
    api_key: str | None = None,
    output_file: str = "youtube_profiles.xlsx",
    max_channels: int = 150,
    fetch_video_stats: bool = True,
    video_stats_mode: str = "full",
):
    api_key = api_key or os.environ.get("YOUTUBE_API_KEY")
    if not api_key:
        sys.exit("ERROR: Set YOUTUBE_API_KEY env var or pass --api-key")

    youtube = get_youtube_client(api_key)

    # 1. Collect channels from keyword searches
    all_channels: dict[str, dict] = {}

    for kw in keywords:
        logger.info("[1/3] Searching '%s' | region=%s | last %dd …", kw, region_code, days)
        found = search_videos_by_keyword(youtube, kw, region_code, days, language, max_channels)
        logger.info("      → %d channels found", len(found))
        merge_keyword_results(all_channels, found)

    if not all_channels:
        logger.warning("No channels found. Try different keywords or a wider date range.")
        return

    channel_ids = list(all_channels.keys())[:max_channels]
    logger.info("[2/3] Fetching details for %d channels …", len(channel_ids))
    channel_details = get_channel_details(youtube, channel_ids)

    # 2. Build profiles
    logger.info("[3/3] Computing metrics …")
    profiles = []
    collected_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Resolve effective mode: legacy fetch_video_stats=False maps to "none"
    effective_mode = video_stats_mode
    if not fetch_video_stats and video_stats_mode == "full":
        effective_mode = "none"

    for cid in tqdm(channel_ids, unit="channel"):
        details = channel_details.get(cid, {})
        search_data = all_channels[cid]

        if effective_mode == "full":
            try:
                vstats = get_recent_video_stats(youtube, cid, days)
                time.sleep(RATE_LIMIT_VIDEO_STATS)
            except HttpError:
                vstats = dict(ZERO_VIDEO_STATS)
        elif effective_mode == "fast":
            vstats = get_video_stats_batch(youtube, search_data.get("video_ids", []))
        else:  # "none"
            vstats = dict(ZERO_VIDEO_STATS)

        has_stats = effective_mode != "none"
        metrics = compute_channel_metrics(details, vstats, search_data, days)
        profile = build_channel_profile(cid, details, search_data, metrics, has_stats, collected_at)
        profiles.append(profile)

    logger.info("%d profiles collected.", len(profiles))
    export_excel(profiles, output_file, keywords)
    return profiles


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    logging.basicConfig(level=logging.INFO, format="%(message)s")

    parser = argparse.ArgumentParser(description="Scrape YouTube profiles by keyword, country, and time period.")
    parser.add_argument(
        "--keywords",
        "-k",
        nargs="+",
        required=True,
        help='One or more keywords, e.g. --keywords "Sorare" "NFT foot"',
    )
    parser.add_argument(
        "--region",
        "-r",
        default="FR",
        help="ISO 3166-1 alpha-2 country code (default: FR)",
    )
    parser.add_argument(
        "--days",
        "-d",
        type=int,
        default=90,
        help="Look-back window in days (default: 90)",
    )
    parser.add_argument(
        "--language",
        "-l",
        default=None,
        help='Relevance language hint, e.g. "fr" (optional)',
    )
    parser.add_argument(
        "--output",
        "-o",
        default="youtube_profiles.xlsx",
        help="Output Excel file name (default: youtube_profiles.xlsx)",
    )
    parser.add_argument(
        "--max-channels",
        "-m",
        type=int,
        default=150,
        help="Max channels to retrieve per keyword (default: 150)",
    )
    parser.add_argument(
        "--api-key",
        default=None,
        help="YouTube Data API v3 key (overrides YOUTUBE_API_KEY env var)",
    )
    parser.add_argument(
        "--no-video-stats",
        action="store_true",
        help="Skip per-channel video stats fetch (backward compat, same as --video-stats-mode none)",
    )
    parser.add_argument(
        "--video-stats-mode",
        choices=["none", "fast", "full"],
        default="full",
        help="Video stats mode: none (skip), fast (reuse search video IDs), full (per-channel search). Default: full",
    )

    args = parser.parse_args()

    # --no-video-stats overrides --video-stats-mode for backward compat
    mode = "none" if args.no_video_stats else args.video_stats_mode

    scrape(
        keywords=args.keywords,
        region_code=args.region.upper(),
        days=args.days,
        language=args.language,
        api_key=args.api_key,
        output_file=args.output,
        max_channels=args.max_channels,
        fetch_video_stats=not args.no_video_stats,
        video_stats_mode=mode,
    )


if __name__ == "__main__":
    main()
